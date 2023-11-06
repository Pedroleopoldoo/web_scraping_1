import openpyxl
from openpyxl.styles import Font, PatternFill


def Colocando_dados():
    import requests
    # Abrindo planilha Excel
    arquivo = openpyxl.load_workbook('Arquivos/projeto_1.xlsx')
    planilha = arquivo['Sheet1']

    # Abrindo api
    url = "https://opcoes.net.br/listaopcoes/completa?au=False&uinhc=0&idLista=ML&idAcao=PETR4&listarVencimentos=true&cotacoes=true"
    pesquisa = requests.get(url).json()

    # Colocando títulos[valores]
    planilha[f'A1'].value = 'column 1'
    planilha[f'B1'].value = 'column 2'
    planilha[f'C1'].value = 'column 3'
    planilha[f'D1'].value = 'column 4'
    planilha[f'E1'].value = 'column 5'
    planilha[f'F1'].value = 'column 6'
    planilha[f'G1'].value = 'column 7'
    planilha[f'H1'].value = 'column 8'
    planilha[f'I1'].value = 'column 9'
    planilha[f'J1'].value = 'column 10'
    planilha[f'K1'].value = 'column 11'
    planilha[f'L1'].value = 'column 12'
    planilha[f'M1'].value = 'column 13'
    planilha[f'N1'].value = 'column 14'
    planilha[f'O1'].value = 'column 15'
    planilha[f'P1'].value = 'column 16'
    planilha[f'Q1'].value = 'column 17'
    planilha[f'R1'].value = 'column 18'

    # Colocando dados
    cont = 1
    for c in pesquisa["data"]["cotacoesOpcoes"]:
        if len(c) == 18:
            cont += 1
        planilha[f'A{cont}'].value = c[0]
        planilha[f'B{cont}'].value = c[1]
        planilha[f'C{cont}'].value = c[2]
        planilha[f'D{cont}'].value = c[3]
        planilha[f'E{cont}'].value = c[4]
        planilha[f'F{cont}'].value = c[5]
        planilha[f'G{cont}'].value = c[6]
        planilha[f'H{cont}'].value = c[7]
        planilha[f'I{cont}'].value = c[8]
        planilha[f'J{cont}'].value = c[9]
        planilha[f'K{cont}'].value = c[10]
        planilha[f'L{cont}'].value = c[11]
        planilha[f'M{cont}'].value = c[12]
        planilha[f'N{cont}'].value = c[13]
        planilha[f'O{cont}'].value = c[14]
        planilha[f'P{cont}'].value = c[15]
        planilha[f'Q{cont}'].value = c[16]
        planilha[f'R{cont}'].value = c[17]

    arquivo.save('Arquivos/projeto_1.xlsx')


def transformando_em_data():
    from datetime import datetime

    arquivo = openpyxl.load_workbook('Arquivos/projeto_1.xlsx')
    planilha = arquivo['Sheet1']

    for c in range(2, len(planilha['A'])):
        if planilha[f"L{c}"].value is not None:
            planilha[f"L{c}"].value = datetime.strptime(planilha[f"L{c}"].value, '%d/%m/%Y')
        else:
            continue
    arquivo.save('Arquivos/projeto_1.xlsx')


def Colocando_em_ordem():
    import pandas as pd

    # Carregar o arquivo Excel
    nome_arquivo = 'Arquivos/projeto_1.xlsx'
    nome_planilha = 'Sheet1'  # Mude para o nome da sua planilha

    # Ler o arquivo Excel
    dados = pd.read_excel(nome_arquivo, sheet_name=nome_planilha)

    # Ordenar os dados pela coluna de datas (coluna B neste caso)
    dados_ordenados = dados.sort_values('column 12', ascending=True)  # Substitua 'Data' pelo nome da sua coluna de datas

    # Salvar os dados ordenados em um novo arquivo Excel
    nome_novo_arquivo = 'dados_ordenados.xlsx'
    dados_ordenados.to_excel('Arquivos/projeto_1.xlsx', index=False)


def arrumando_colunas():

    # Abrindo planilha Excel
    arquivo = openpyxl.load_workbook('Arquivos/projeto_1.xlsx')
    planilha = arquivo['Sheet1']

    # Titulos[formato]
    planilha[f'A1'].font = Font(color='FFFFFF', size=16)
    planilha[f'B1'].font = Font(color='FFFFFF', size=16)
    planilha[f'C1'].font = Font(color='FFFFFF', size=16)
    planilha[f'D1'].font = Font(color='FFFFFF', size=16)
    planilha[f'E1'].font = Font(color='FFFFFF', size=16)
    planilha[f'F1'].font = Font(color='FFFFFF', size=16)
    planilha[f'G1'].font = Font(color='FFFFFF', size=16)
    planilha[f'H1'].font = Font(color='FFFFFF', size=16)
    planilha[f'I1'].font = Font(color='FFFFFF', size=16)
    planilha[f'J1'].font = Font(color='FFFFFF', size=16)
    planilha[f'K1'].font = Font(color='FFFFFF', size=16)
    planilha[f'L1'].font = Font(color='FFFFFF', size=16)
    planilha[f'M1'].font = Font(color='FFFFFF', size=16)
    planilha[f'N1'].font = Font(color='FFFFFF', size=16)
    planilha[f'O1'].font = Font(color='FFFFFF', size=16)
    planilha[f'P1'].font = Font(color='FFFFFF', size=16)
    planilha[f'Q1'].font = Font(color='FFFFFF', size=16)
    planilha[f'R1'].font = Font(color='FFFFFF', size=16)

    planilha[f'A1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'B1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'C1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'D1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'E1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'F1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'G1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'H1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'I1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'J1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'K1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'L1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'M1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'N1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'O1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'P1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'Q1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    planilha[f'R1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    planilha.column_dimensions['A'].width = 20
    planilha.column_dimensions['B'].width = 15
    planilha.column_dimensions['C'].width = 15
    planilha.column_dimensions['D'].width = 15
    planilha.column_dimensions['E'].width = 15
    planilha.column_dimensions['F'].width = 15
    planilha.column_dimensions['G'].width = 15
    planilha.column_dimensions['H'].width = 15
    planilha.column_dimensions['I'].width = 15
    planilha.column_dimensions['J'].width = 15
    planilha.column_dimensions['K'].width = 15
    planilha.column_dimensions['L'].width = 20
    planilha.column_dimensions['M'].width = 31
    planilha.column_dimensions['N'].width = 31
    planilha.column_dimensions['O'].width = 31
    planilha.column_dimensions['P'].width = 31
    planilha.column_dimensions['Q'].width = 31
    planilha.column_dimensions['R'].width = 15

    arquivo.save('Arquivos/projeto_1.xlsx')


def executando():
    Colocando_dados()
    transformando_em_data()
    Colocando_em_ordem()
    arrumando_colunas()


if __name__ == '__main__':
    executando()
